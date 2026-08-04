"""
메리츠 수입(EV) 장기렌트 엑셀 → JSON 추출기 — 테슬라·폴스타·BYD.

`차량정보` 시트는 국산 렌트와 같은 "브랜드별로 나란히 놓인 여러 블록" 구조다
(일반=TESLA와 완전 동일 중복, TESLA, 폴스타, BYD) — extract_meritz_rental_domestic.py
의 find_blocks() 패턴을 그대로 재사용해 모든 블록을 순회한다. 다만 첫 두 블록
(E열·BV열)이 완전히 동일한 테슬라 중복이라, 먼저 나온 블록만 남기고 이후
중복 모델명은 스킵한다.

Model 3 RWD(보조금) 실제 셀 값(견적조건 시트, 기본가격 45,000,000원 · 전기차
보조금 2,200,000원)으로 취득원가 계산 사슬을 셀 단위까지 검증했다(오차 0):

  면세가격A = ROUND((기본가격+옵션)/개소세계수 + 1차탁송료/1.1, 0) × 1.1(VAT 포함)
  출고가격  = 면세가격A − 할인
  공급가격B = ROUND(출고가격/1.1, 0)
  기준세액(A) = 공급가격B × 5% × 130%   ← 국산(4%)과 다르다, 수입 EV는 5%
  개별소비세 = MAX(기준세액A − EV감면(390만), 0)  (EV는 거의 항상 0으로 수렴)
  취득세    = TRUNC(공급가격B×4% − EV취득세감면(140만, 단 잔가군열이 "X"면 미적용), -10)
  취득원가  = 공급가격B + (2차탁송료+용품비)/1.1 + 개별소비세 + 취득세
             + 등록제비용 − 전기차보조금(EV만, 수기입력값)
  → 취득원가 39,185,451원 (정확히 일치, 견적조건 시트 수식 셀 단위 확인)

⚠ 전기차보조금은 지역·시기별로 바뀌는 수기입력값(렌트_입력!X15)이라 카탈로그에
  없다 — 엔진 입력으로 받되 기본값 0(보조금을 모른다고 견적이 실제보다 낮게
  보이면 안 되므로, 모르면 0원으로 두는 게 안전한 방향).
⚠ `정비` 시트는 이 파일에서 전 모델·전 등급(Platinum/Standard/Select)이
  0원이라 정비비는 항상 0으로 취급한다(국산 렌트의 Basic 단가표와 다른 점,
  별도 maintenance.json 없음).
⚠ 금리 등급표(전략AA=6.5% 등)도 국산과 다른 별도 표 — 절대 재사용하지 않는다
  (lib/engine/meritz-rental-import/index.ts의 STRATEGY_RATES 참고, 견적조건
  시트 F35:H50 구간에서 셀 값 그대로 옮김).

사용법: python scripts/extract_meritz_rental_import.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/meritz-rental-import/data"

TERM_MONTHS = [24, 36, 48, 60]
MILEAGE_BUCKETS = [10000, 15000, 20000, 25000, 30000, 40000, None]  # None=무제한
RESIDUAL_START_COL = 15  # 모델명 컬럼 기준 상대 오프셋(국산 스크립트와 동일 레이아웃)


def find_blocks(ws) -> list[tuple[int, int]]:
    blocks: list[tuple[int, int]] = []
    for r in range(1, 10):
        for c in range(1, 600):
            if ws.cell(row=r, column=c).value == "차종":
                blocks.append((r, c))
    return blocks


def extract_vehicles(ws) -> dict:
    blocks = find_blocks(ws)
    if not blocks:
        raise RuntimeError("차량정보 헤더('차종') 못 찾음")

    vehicles: dict[str, dict] = {}
    skipped_dupe = 0
    for hdr_row, hdr_col in blocks:
        for r in range(hdr_row + 1, hdr_row + 200):
            model = ws.cell(row=r, column=hdr_col).value
            if not model or not isinstance(model, str):
                continue
            model = model.strip()
            if model.startswith("■") or model in ("", "-", "차종") or set(model) in ({"-"}, {"ㅡ"}):
                continue

            tax_factor = ws.cell(row=r, column=hdr_col + 1).value  # 개소세계수
            kind = ws.cell(row=r, column=hdr_col + 2).value  # 차량등급
            insurance_grade = ws.cell(row=r, column=hdr_col + 3).value  # 보험등급
            strategy_grade = ws.cell(row=r, column=hdr_col + 4).value  # 전략구분
            fuel = ws.cell(row=r, column=hdr_col + 5).value  # 유종
            residual_group_flag = ws.cell(row=r, column=hdr_col + 9).value  # 잔가군("X"=EV 취득세감면 제외)

            residual: dict[str, float] = {}
            col = hdr_col + RESIDUAL_START_COL - 5
            for term in TERM_MONTHS:
                for mileage in MILEAGE_BUCKETS:
                    v = ws.cell(row=r, column=col).value
                    if isinstance(v, (int, float)) and mileage is not None and v > 0:
                        residual[f"{term}_{mileage}"] = round(float(v), 4)
                    col += 1

            if not residual:
                continue

            if model in vehicles:
                skipped_dupe += 1
                continue

            vehicles[model] = {
                "fuel": fuel or "",
                "kind": kind or "",
                "insuranceGrade": insurance_grade or "",
                "strategyGrade": strategy_grade or "일반",
                "consumptionTaxFactor": round(float(tax_factor), 6)
                if isinstance(tax_factor, (int, float)) and tax_factor
                else 1.1,
                "evAcquisitionTaxRebateExcluded": residual_group_flag == "X",
                "residualByTermMileage": residual,
            }
    if skipped_dupe:
        print(f"  (중복 모델명 {skipped_dupe}건은 먼저 나온 블록 값 유지)")
    return vehicles


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    vehicles = extract_vehicles(wb["차량정보"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(vehicles, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 모델")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_meritz_rental_import.py <엑셀파일>")
        sys.exit(1)
    main(sys.argv[1])
