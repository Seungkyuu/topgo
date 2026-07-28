"""
메리츠캐피탈 렌터카(장기렌트) 엑셀 → JSON 추출기 (2단계)

─── 사용법 ────────────────────────────────────────────────────────────────────

1단계 — 구조 파악:
  python scripts/extract_meritz.py discover <엑셀파일>

2단계 — 실제 추출 (EXTRACT_CONFIG 채운 뒤):
  python scripts/extract_meritz.py extract <엑셀파일>

─── 골든케이스 ─────────────────────────────────────────────────────────────────
  차량가 45,000,000 / 36개월 / 잔가율 57% / 금리 6%
  원금 39,431,000 / 잔가 25,650,000
  자동차세 2,000 / 보험 58,400 / 정비 106,000
  → 매회납부 851,840원 ✅
"""

import sys
import json
import math
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl 후 재실행")
    sys.exit(1)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2단계에서 채울 설정 (discover 결과 보고 수정)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXTRACT_CONFIG = {
    "rental_sheet": None,       # 예: "렌트_입력"
    "condition_sheet": None,    # 예: "견적조건"

    # 모델/원금 테이블
    "model_col":       None,    # 모델명 컬럼
    "principal_col":   None,    # 원금 컬럼
    "table_row_start": None,
    "table_row_end":   None,

    # 잔가율 테이블 (렌터카용)
    "res_model_col":   None,
    "res_36_col":      None,
    "res_48_col":      None,
    "res_60_col":      None,
    "res_row_start":   None,
    "res_row_end":     None,

    # 보험·정비 테이블
    "ins_model_col":   None,    # 보험료 모델컬럼
    "ins_36_col":      None,    # 36개월 보험
    "ins_48_col":      None,
    "ins_60_col":      None,
    "maint_col":       None,    # 정비비 (단일 또는 기간별)
}

OUT_DIR = Path(__file__).parent.parent / "lib/engine/meritz/data"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 수식 유틸 (골든케이스 검증용)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def pmt(rate, nper, pv, fv=0):
    if rate == 0:
        return -(pv + fv) / nper
    f = (1 + rate) ** nper
    return -(pv * f + fv) * rate / (f - 1)

def roundup(v, digits):
    factor = 10 ** digits
    return (math.ceil(v / factor) if v >= 0 else -math.ceil(-v / factor)) * factor

def rounddown(v, digits):
    factor = 10 ** digits
    return (math.floor(v / factor) if v >= 0 else -math.floor(-v / factor)) * factor

def verify_golden():
    rate = 0.06 / 12
    pv = -39_431_000
    fv = round(25_650_000 / 1.1) - 0
    finance = roundup(pmt(rate, 36, pv, fv), -2)
    supply = finance + 2_000 + 58_400 + 106_000 + 500 + 700
    vat = rounddown(supply * 0.1, 0)
    monthly = supply + vat
    ok = monthly == 851_840
    print(f"  골든케이스: {monthly:,}원  {'✅' if ok else '❌ 예상 851,840'}")
    return ok


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1단계: discover
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def discover(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("=" * 60)
    print(f"파일: {Path(xlsx_path).name}")
    print(f"시트 목록: {wb.sheetnames}")
    print("=" * 60)

    print("\n[골든케이스 수식 검증]")
    verify_golden()

    for sname in wb.sheetnames:
        ws = wb[sname]
        count = sum(1 for row in ws.iter_rows() for c in row if c.value is not None)
        print(f"\n▶ 시트 [{sname}] — {count}개 비어있지 않은 셀")
        # 처음 40행 × A~BJ
        printed = 0
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=62):
            for cell in row:
                v = cell.value
                if v is not None and str(v).strip():
                    print(f"  {cell.coordinate}: {str(v)[:60]}")
                    printed += 1
        if printed == 0:
            print("  (첫 40행에 데이터 없음)")

    print("\n" + "=" * 60)
    print("위 결과를 확인하고 EXTRACT_CONFIG를 채운 뒤 extract 명령 실행")
    print("=" * 60)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2단계: extract
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def extract(xlsx_path: str):
    cfg = EXTRACT_CONFIG
    required = ["rental_sheet", "model_col", "principal_col",
                "table_row_start", "table_row_end"]
    missing = [k for k in required if cfg[k] is None]
    if missing:
        print(f"EXTRACT_CONFIG 미완성: {missing}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[cfg["rental_sheet"]]

    vehicles = {}
    for row in range(cfg["table_row_start"], cfg["table_row_end"] + 1):
        model = ws[f"{cfg['model_col']}{row}"].value
        principal = ws[f"{cfg['principal_col']}{row}"].value
        if not model or not principal:
            continue
        vehicles[str(model).strip()] = {"principal": int(principal)}

    residual_rates = {}
    if cfg["res_model_col"]:
        for row in range(cfg["res_row_start"], cfg["res_row_end"] + 1):
            model = ws[f"{cfg['res_model_col']}{row}"].value
            if not model:
                continue
            residual_rates[str(model).strip()] = {
                "36": float(ws[f"{cfg['res_36_col']}{row}"].value or 0),
                "48": float(ws[f"{cfg['res_48_col']}{row}"].value or 0),
                "60": float(ws[f"{cfg['res_60_col']}{row}"].value or 0),
            }

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    (OUT_DIR / "vehicles.json").write_text(
        json.dumps({"_note": "메리츠 렌터카 엑셀 추출", **vehicles},
                   ensure_ascii=False, indent=2)
    )
    print(f"✅ vehicles.json ({len(vehicles)}개 모델)")

    if residual_rates:
        (OUT_DIR / "residual-rates.json").write_text(
            json.dumps({"_note": "메리츠 렌트 잔가율표", **residual_rates},
                       ensure_ascii=False, indent=2)
        )
        print(f"✅ residual-rates.json ({len(residual_rates)}개 모델)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    cmd, path = sys.argv[1], sys.argv[2]
    if cmd == "discover":
        discover(path)
    elif cmd == "extract":
        extract(path)
    else:
        print(f"알 수 없는 명령: {cmd}  (discover | extract)")
        sys.exit(1)
