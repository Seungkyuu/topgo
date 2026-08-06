"""
메리츠 테슬라 전용 운용리스 엑셀(2608_V2) → JSON 추출기.

`차량정보` 시트 구조는 국산/수입 렌트 추출기와 같은 "블록" 패턴이지만 컬럼
배치가 다르다(브랜드가 구분행이 아니라 매 행의 "제조사" 열에 직접 들어있다):

  헤더행(hdr_row)의 "차종" 셀을 hdr_col이라 하면,
    hdr_col-1 = 제조사(브랜드)   hdr_col+1 = 종류   hdr_col+3 = 연료
    hdr_col+5 = 배기량           hdr_col+6 = 차세(연간 자동차세)
    hdr_col+7 = 잔가율표 시작(24개월 10,000km) — 이후 6버킷×4구간 이어짐

차종 블록이 2개 있는데(hdr_col=8, hdr_col=53) 완전히 동일한 테슬라 중복이라
먼저 나온 블록만 남기고 이후 같은 모델명은 스킵한다(extract_meritz_rental_
import.py와 동일 원칙).

산출: lib/engine/meritz-tesla/data/vehicles.json — meritz-domestic/meritz-
rental-domestic과 동일한 스키마(brand/kind/fuel/engineCc/annualTax/
residualByTermMileage), lib/engine/meritz-ev/lease.ts(LeaseformVehicle)가
그대로 소비한다.

사용법: python scripts/extract_meritz_tesla_lease.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/meritz-tesla/data"

TERM_MONTHS = [24, 36, 48, 60]
MILEAGE_BUCKETS = [10000, 15000, 20000, 25000, 30000, 40000]
RESIDUAL_START_OFFSET = 7  # hdr_col + 7 = 첫 잔가율 컬럼


def find_blocks(ws) -> list[tuple[int, int]]:
    blocks: list[tuple[int, int]] = []
    for r in range(1, 10):
        for c in range(1, 200):
            if ws.cell(row=r, column=c).value == "차종":
                blocks.append((r, c))
    return blocks


def extract_vehicles(ws) -> dict:
    blocks = find_blocks(ws)
    if not blocks:
        raise RuntimeError("차량정보 헤더('차종') 못 찾음")

    vehicles: dict[str, dict] = {}
    for hdr_row, hdr_col in blocks:
        for r in range(hdr_row + 1, hdr_row + 200):
            model = ws.cell(row=r, column=hdr_col).value
            if not model or not isinstance(model, str):
                continue
            model = model.strip()
            if model in vehicles:
                continue  # 중복 블록(테슬라 2블록 완전 동일) — 먼저 나온 값 우선
            if model in ("", "-", "차종") or set(model) <= set("■―ㅡ-ー—"):
                continue

            brand = ws.cell(row=r, column=hdr_col - 1).value or "테슬라"
            kind = ws.cell(row=r, column=hdr_col + 1).value
            fuel = ws.cell(row=r, column=hdr_col + 3).value
            engine_cc = ws.cell(row=r, column=hdr_col + 5).value or 0
            annual_tax = ws.cell(row=r, column=hdr_col + 6).value or 0

            residual: dict[str, float] = {}
            col = hdr_col + RESIDUAL_START_OFFSET
            for term in TERM_MONTHS:
                for mileage in MILEAGE_BUCKETS:
                    v = ws.cell(row=r, column=col).value
                    if isinstance(v, (int, float)):
                        residual[f"{term}_{mileage}"] = round(float(v), 4)
                    col += 1

            if not residual or not any(residual.values()):
                continue  # 구분행 등 잔가율 전부 0/미기입인 비차량 행

            vehicles[model] = {
                "brand": str(brand).strip(),
                "kind": str(kind or ""),
                "fuel": str(fuel or ""),
                "engineCc": int(engine_cc) if isinstance(engine_cc, (int, float)) else 0,
                "annualTax": int(annual_tax) if isinstance(annual_tax, (int, float)) else 0,
                "residualByTermMileage": residual,
            }
    return vehicles


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    vehicles = extract_vehicles(wb["차량정보"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(vehicles, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 모델")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_meritz_tesla_lease.py <엑셀파일>")
        sys.exit(1)
    main(sys.argv[1])
