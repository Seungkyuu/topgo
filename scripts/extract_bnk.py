"""
BNK캐피탈 운용리스 엑셀 → JSON 추출기.

이 엑셀은 다른 캐피탈과 근본적으로 다른 구조다 — 잔가율이 모델별 고정표가
아니라 **7개 잔가보증사(WS/SE(CB)/BR/TY/JY/CR/ADB) 중 최고가를 자동 선택**하는
구조다. 각 잔가사는 차종마다 자기만의 "잔가군 코드"를 가질 수도, 없을 수도
있다(취급 안 하면 코드가 비어있음) — 그 코드로 RVs 시트의 "공유" 잔가율
매트릭스(개월×잔가군, 전 잔가사 공용)를 조회한다.

Es1 시트에서 셀 단위로 검증한 사슬(BMW 520i, JY 잔가사 자동선택 예시):

  등록세 = ROUNDDOWN(차량가/1.1 × 5%, -10)                      = 5,000,000
  취득세 = ROUNDDOWN(차량가/1.1 × 2%, -10)  [비EV·비경차 표준]    = 2,000,000
  취득원가 = 차량가 + 등록세 + 취득세 (+탁송 등, 골든케이스는 0)  = 117,000,000
  잔가사별 기본잔가율 = RVs!AG8:CW67에서 INDEX/MATCH(개월, 잔가군코드)
  최대잔가 = 기본잔가율 + 주행거리감가 + 최대인상폭(전 잔가사 공통 7%)
  실적용 잔가사 = 7개 잔가사 중 최대잔가가 가장 높은 곳 (여기선 JY, 0.625)
  잔가 = ROUNDUP(차량가 × 잔가율, -3)                            = 68,750,000
  잔가보장수수료 = 잔가사별 수수료율(공유 브래킷표, 상단 티어만 잔가사 개별) × 취득원가
  기본리스료 = ROUNDUP(PMT(IRR/12, 개월, -취득원가-수수료, 잔가), -1) = 1,260,000

⚠ 이번 v1에서 의도적으로 뺀 것(전부 "실제보다 낮게 보이면 안 된다" 원칙으로
  안전한 방향 — 나중에 상담에서 조건이 더 좋아질 수 있는 건 괜찮다):
  · 딜러사별 제휴 IRR 우대(Cond 시트, 예: BMW-동성모터스 5.41%) — 딜러사를
    묻지 않는 우리 UI 특성상 "비제휴" 최고금리(7.41%)를 항상 사용한다.
  · 특판/프로모션 잔가 상향, 사전협의 잔가, 차종지원금 — 전부 미반영(0).
  · EV/경차/승합/화물 특례 취득세·등록세율 — 표준(승용·비EV) 세율만 반영.
    친환경차는 카탈로그에 '친환경차'=1로 표시되지만, 세율 특례가 아직
    미검증이라 일단 표준 세율로 근사(실제보다 비싸게 보일 수 있음=안전).

사용법: python scripts/extract_bnk.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/bnk/data"

# CDB 시트 컬럼(1-based) — row3 헤더로 확인
COL = {
    "국산수입": 3,
    "DB코드": 4,
    "브랜드": 5,
    "차명세부": 7,
    "년식": 8,
    "미션": 9,
    "차량명": 10,
    "차량등급": 13,
    "배기량": 14,
    "친환경차": 15,
    "WS": 11,
    "CB": 16,
    "BR": 17,
    "TY": 18,
    "JY": 19,
    "CR": 20,
    "ADB": 36,
}
GUARANTORS = ["WS", "CB", "BR", "TY", "JY", "CR", "ADB"]


def extract_vehicles(ws) -> dict:
    vehicles: dict[str, dict] = {}
    for r in range(4, ws.max_row + 1):
        code = ws.cell(row=r, column=COL["DB코드"]).value
        name = ws.cell(row=r, column=COL["차량명"]).value
        if not code or not name:
            continue
        guarantor_codes = {}
        for g in GUARANTORS:
            v = ws.cell(row=r, column=COL[g]).value
            if v is not None and v != 0:
                guarantor_codes[g] = str(v)
        if not guarantor_codes:
            continue  # 어느 잔가사도 취급 안 함 — 견적 불가 차종

        vehicles[str(code)] = {
            "brand": ws.cell(row=r, column=COL["브랜드"]).value or "",
            "model": name,
            "isImport": ws.cell(row=r, column=COL["국산수입"]).value == "수입",
            "kind": ws.cell(row=r, column=COL["차량등급"]).value or "",
            "engineCc": ws.cell(row=r, column=COL["배기량"]).value or 0,
            "isEco": bool(ws.cell(row=r, column=COL["친환경차"]).value),
            "guarantorCodes": guarantor_codes,
        }
    return vehicles


def extract_residual_matrix(ws) -> dict:
    """RVs!AF8:CW67 — 개월(1~60) × 잔가군코드 → 기본잔가율. 전 잔가사 공용."""
    class_cols = list(range(33, 102))  # AG~CW
    classes = [ws.cell(row=7, column=c).value for c in class_cols]
    matrix: dict[str, dict[str, float]] = {}
    for r in range(8, 68):
        month = ws.cell(row=r, column=32).value  # AF열
        if month is None:
            continue
        row_vals: dict[str, float] = {}
        for c, cls in zip(class_cols, classes):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and cls is not None:
                row_vals[str(cls)] = round(float(v), 6)
        matrix[str(int(month))] = row_vals
    return matrix


def extract_fee_table(ws) -> dict:
    """Es1!F168:J175 — 공유 브래킷(하향율 임계값 → 중간 티어 수수료율) +
    잔가사별 최상단(>6~7%) 티어 수수료율."""
    thresholds = []
    shared_rates = []
    for r in range(168, 176):
        thr = ws.cell(row=r, column=6).value
        rate = ws.cell(row=r, column=8).value
        if isinstance(thr, (int, float)) and isinstance(rate, (int, float)):
            thresholds.append(round(float(thr), 4))
            shared_rates.append(round(float(rate), 6))
    top_tier: dict[str, float] = {}
    for r in range(168, 175):
        name = ws.cell(row=r, column=9).value
        rate = ws.cell(row=r, column=10).value
        if name and isinstance(rate, (int, float)):
            top_tier[str(name)] = round(float(rate), 6)
    return {
        "thresholds": thresholds,  # 내림차순, 각 임계값 초과분에 sharedRates[i] 적용
        "sharedRates": shared_rates,
        "topTierByGuarantor": top_tier,  # 잔가사가 "하향 없이"(>6~7% 구간) 받는 우대 수수료율
    }


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    vehicles = extract_vehicles(wb["CDB"])
    residual_matrix = extract_residual_matrix(wb["RVs"])
    fee_table = extract_fee_table(wb["Es1"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(vehicles, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    (OUT_DIR / "residual-matrix.json").write_text(
        json.dumps(residual_matrix, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    (OUT_DIR / "fee-table.json").write_text(
        json.dumps(fee_table, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 모델")
    print(f"✅ residual-matrix.json: {len(residual_matrix)}개월 구간")
    print(f"✅ fee-table.json: {fee_table}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_bnk.py <엑셀파일>")
        sys.exit(1)
    main(sys.argv[1])
