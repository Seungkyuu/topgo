"""
메리츠 Polestar 전용 운용리스 엑셀(2608_V2) → JSON 추출기.

일반 수입차 리스(extract_meritz_lease.py)와 별개로 존재하는 Polestar 4
생산배치별(1차시/2차시/3차시) 세부 트림 전용 견적 시트다. `차종` 시트에는
차량가가 없다(0) — "운용리스 내부" 시트에 살아있는 유일한 계산 예시
(Long Range Dual Motor + PLDC 3차시, 차량가 70,000,000 → 할인 후
58,700,000)로 아래 사슬을 셀 단위까지 검증했다(오차 0):

  차량가최종 = 차량가 − (차량가×할인율 + 추가할인)
  과세표준   = 차량가최종 / 1.1
  등취득세   = ROUNDDOWN(과세표준×7%, -1) − 140만(EV 취득세감면, `취득세`시트
               "전기" 카테고리 공채 0·감면 그대로 확인)
  취득원가   = 차량가최종 + 등취득세 + 공채(EV는 0) + 탁송료 + 부대비 + 추가부대비
  잔가율     = {36:0.65, 48:0.6, 60:0.55} — "잔가" 시트에 "자체잔가" 단일
               잔가사만 쓰는 게 확인됨(다른 잔가사 컬럼은 전부 공란).
               ⚠ 이 잔가율표는 이 골든케이스 트림 기준으로 확인된 값이라,
               `차종` 시트에서 전 트림이 같은 잔가군("웨스트" 컬럼="폴스타")
               을 공유하는 걸 근거로 전 트림에 동일 적용한다 — 트림별로
               실제 다르다면 대표님이 실제 견적으로 알려주시면 바로잡는다.
  잔가       = ROUNDDOWN(차량가최종×잔가율, -3)
  CM/AG 수수료·잔가보장수수료 = 0 (골든케이스에서 전부 0으로 확인 — "CM/AG
               최대 7%"는 상한 메모일 뿐 이 트림엔 적용 안 됨. 다른 트림·
               조건에서 수수료가 붙는지는 미확인 — 안전한 기본값 0으로 시작)
  리스료     = ROUNDUP(PMT(IRR/12, 기간, −(취득원가−보증금−선납금+수수료),
               잔가−보증금−선납금), -2)

사용법: python scripts/extract_meritz_polestar_lease.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/meritz-polestar/data"

RESIDUAL_RATE_BY_TERM = {"36": 0.65, "48": 0.6, "60": 0.55}


def extract(path: str) -> None:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["차종"]

    vehicles: dict[str, dict] = {}
    for r in range(7, 40):
        maker = ws.cell(row=r, column=2).value
        trim = ws.cell(row=r, column=5).value
        if not maker or not trim or not isinstance(trim, str):
            continue
        trim = trim.strip()
        if trim in ("", "-"):
            continue
        model2 = ws.cell(row=r, column=4).value  # "Polestar4" 등 기본모델
        fuel = ws.cell(row=r, column=7).value
        kind = ws.cell(row=r, column=8).value
        ev_flag = ws.cell(row=r, column=9).value
        discount_rate = ws.cell(row=r, column=16).value
        discount_amount = ws.cell(row=r, column=17).value

        display = f"{str(model2).strip()} {trim}" if model2 else trim
        vehicles[display] = {
            "brand": str(maker).strip(),
            "kind": str(kind or ""),
            "fuel": str(fuel or ""),
            "evAcquisitionTaxRebate": ev_flag == "전기",
            "discountRate": float(discount_rate or 0),
            "discountAmount": int(discount_amount or 0),
            "residualByTerm": dict(RESIDUAL_RATE_BY_TERM),
        }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    v = dict(vehicles)
    v["_note"] = "메리츠 Polestar 전용 운용리스 2608_V2 '차종' 시트에서 추출"
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(v, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 트림")
    for k in vehicles:
        print(" -", k)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_meritz_polestar_lease.py <엑셀파일>")
        sys.exit(1)
    extract(sys.argv[1])
