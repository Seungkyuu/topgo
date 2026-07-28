"""
메리츠캐피탈 운용리스·금융리스 엑셀(MeritzCapital 2607_V1) → JSON 추출기

사용법:
  python scripts/extract_meritz_lease.py extract <엑셀파일(.xlsm)>

─── 운용리스 수식 체인 ('운용리스 내부' 시트 해부) ───────────────────────────────
  취득원가(H8)   = 차량가최종 + 등취득세 + 공채 + 탁송료 + 부대비
    등취득세     = ROUNDDOWN(차량가최종/1.1 × 7%, -1)   [감면 미반영 D18과 일치]
  보증금(I10)   = ROUNDDOWN(차량가최종 × 보증금율, -3)
  장기선수금(I11)= ROUNDDOWN(차량가최종 × 선수금율, -3)
  잔가(I12)     = ROUNDDOWN(차량가최종 × 잔가율, -3)
  금리(H36)     = 브랜드기준금리 + 저보증금(<11%: +0.15%) + 보증금가산(40%+: +0.3%)
                  + 24MY(+0.5%) 등
  수수료: CM = RD(취득원가×3%,-1), 제휴사 = RD(취득원가×제휴사율,-1),
         추가 = RD(취득원가×1%,-1), 고정 10,000
  잔가보장수수료: 잔가사(West/AJ/APS/VGS/자체) 중 계약잔가 달성 가능한 곳의
         구간별 수수료 (구간 = 계약잔가 - 기본잔가, 1%p당 1구간)
  리스료(H52)   = ROUNDUP(PMT(금리/12, 기간,
                    -취득원가 + 보증금 + 선수금 - CM - 제휴사 - 추가 - 잔가보장 - 10000,
                    잔가 - 보증금), -2)
  고객 월납입   = H52 (+차세 포함 시 가산). 표시리스료 = RD(H52+선수금/기간,-1)

─── 잔가 산정 ('잔가' 시트) ─────────────────────────────────────────────────────
  잔가사별 기본잔가 = 매트릭스[잔가군][기간] + 주행거리조정(3만km: -4%, 자체 -15%)
  고잔가 = 기본 + 8%(West/AJ/APS) / +6%(VGS)  [기간 12~60 & 주행 2만km 이하]
  AJ·자체는 주행 2만 미만 +col13, 1.5만 미만 +col14 추가
  최대잔가 = max(잔가사별 고잔가) / 최소잔가 = {12:50%,24:30%,36:30%,48:20%,60:15%}

─── 금융리스 ('금융리스 내부') ──────────────────────────────────────────────────
  리스료 = ROUNDUP(PMT(금리/12, 기간, -(차량가최종-선수금), 유예금), -1)
  금리는 입력값 (2607 저장분: 6.3%)

─── 골든케이스 (시트 저장값) ────────────────────────────────────────────────────
  운용: Benz E 220d 4MATIC AMG Line / 차량가 81,000,000 / 60개월 / 2만km
        보증금 0 / 선수 10% / 잔가 APS 0.575 → 월 976,700
  금융: BYD DOLPHIN / 65,000,000 / 36개월 / 선수 20% / 6.3% → 월 1,589,020
"""

import sys
import json
import math
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/meritz/data"

PROVIDERS = {
    # (표시명, 매트릭스 시작행(코드행), 기간행수, 기간목록, 차종시트 잔가군 컬럼(E=1 기준), 고잔가 가산)
    "west": (48, [12, 24, 36, 48, 60, 72], 6, 0.08),
    "aj":   (57, [12, 24, 36, 48, 60], 7, 0.08),
    "aps":  (65, [12, 24, 36, 48, 60], 8, 0.08),
    "vgs":  (73, [24, 36, 48, 60], 9, 0.06),
    "self": (80, [24, 36, 48, 60], 10, 0.0),
}
# 수수료표: 잔가!C38:H45 — 구간 1~8, 컬럼 D=W E=J F=A G=V H=자체
FEE_COLS = {"west": 4, "aj": 5, "aps": 6, "vgs": 7, "self": 8}


def pmt(rate, nper, pv, fv=0):
    f = (1 + rate) ** nper
    return -(pv * f + fv) * rate / (f - 1)


def ru(v, d):
    f = 10 ** (-d)
    return math.ceil(v / f) * f


def rd(v, d):
    f = 10 ** (-d)
    return math.floor(v / f) * f


def golden_check():
    """운용리스 골든케이스: E 220d / 60개월 / 선수 10% / 잔가 0.575 → 976,700"""
    price = 81_000_000
    tax = rd(price / 1.1 * 0.07, -1)
    assert tax == 5_154_540, tax
    acq = price + tax + 0 + 0 + 80_000
    assert acq == 86_234_540, acq
    prepay = rd(price * 0.1, -3)
    assert prepay == 8_100_000
    residual = rd(price * 0.575, -3)
    assert residual == 46_575_000
    cm = rd(acq * 0.03, -1)
    assert cm == 2_587_030, cm
    partner = rd(acq * 0.011, -1)
    assert partner == 948_570, partner
    extra = rd(acq * 0.01, -1)
    assert extra == 862_340, extra
    guarantee = 1_053_000  # APS 8구간
    rate = 0.0635 + 0.0015  # Benz + 저보증금(<11%)
    pv = -acq + 0 + prepay - cm - 0 - guarantee - partner - extra - 10_000
    monthly = ru(pmt(rate / 12, 60, pv, residual), -2)
    assert monthly == 976_700, monthly
    # 금융리스: BYD DOLPHIN
    fin = ru(pmt(0.063 / 12, 36, -(65_000_000 - 13_000_000), 0), -1)
    assert fin == 1_589_020, fin
    print("✅ 골든케이스 수식 체인 검증 통과 (운용 976,700 / 금융 1,589,020)")


def extract(path):
    golden_check()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # 1) 차종 마스터
    ws = wb["차종"]
    brands_ws = wb["브랜드"]
    brand_names = [
        str(brands_ws.cell(row=r, column=2).value)
        for r in range(2, 32)
        if brands_ws.cell(row=r, column=2).value
    ]
    vehicles = {}
    for row in ws.iter_rows(min_row=7, max_row=1209, min_col=4, max_col=18, values_only=True):
        d, model, price = row[0], row[1], row[2]
        if not model or not price:
            continue
        brand = next((b for b in brand_names if str(d).startswith(b)), str(d))
        key = f"{brand} {str(model).strip()}"
        if key in vehicles:
            continue  # 첫 등장 우선
        vehicles[key] = {
            "brand": brand,
            "model": str(model).strip(),
            "vehiclePrice": int(price),
            "engineCc": int(row[3] or 0),
            "fuel": str(row[4] or ""),
            "groups": {
                p: str(row[col]).strip()
                for p, (_, _, col, _) in PROVIDERS.items()
                if row[col] is not None
            },
            "highResidualBan": str(row[12] or "") == "고잔가불가",
            "hrBonus15k": float(row[13] or 0),
            "hrBonus10k": float(row[14] or 0),
        }
    print(f"차종 {len(vehicles)}개")

    # 2) 잔가사별 매트릭스
    wsj = wb["잔가"]
    matrices = {}
    for p, (start, terms, _, hr) in PROVIDERS.items():
        codes = []
        col = 3
        while True:
            v = wsj.cell(row=start, column=col).value
            if v is None:
                break
            codes.append(str(v).strip())
            col += 1
        m = {}
        for ci, code in enumerate(codes):
            entry = {}
            for ti, t in enumerate(terms):
                v = wsj.cell(row=start + 1 + ti, column=3 + ci).value
                if v is not None:
                    entry[str(t)] = round(float(v), 4)
            if code in m:
                continue  # 중복 코드(K/M 등)는 첫 컬럼 우선
            m[code] = entry
        matrices[p] = {"rates": m, "highResidualBonus": hr}
        print(f"  {p}: 잔가군 {len(m)}개")

    # 3) 수수료표 (구간 1~8)
    fees = {}
    for p, col in FEE_COLS.items():
        f = {}
        for band in range(1, 9):
            v = wsj.cell(row=37 + band, column=col).value
            f[str(band)] = int(v) if v is not None else None
        fees[p] = f

    # 4) 최소잔가 (J41:K45)
    floors = {}
    for r in range(41, 46):
        t, v = wsj.cell(row=r, column=10).value, wsj.cell(row=r, column=11).value
        if t:
            floors[str(int(t))] = float(v)

    # 5) 브랜드 금리 + 가산
    brand_rates = {}
    for r in range(2, 32):
        b, v = brands_ws.cell(row=r, column=2).value, brands_ws.cell(row=r, column=3).value
        if b and v:
            brand_rates[str(b)] = float(v)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    v = dict(vehicles)
    v["_note"] = "메리츠캐피탈 수입리스 2607_V1 '차종' 시트에서 추출"
    (OUT_DIR / "vehicles.json").write_text(json.dumps(v, ensure_ascii=False, indent=1))
    data = {
        "_note": "메리츠캐피탈 수입리스 2607_V1 — 잔가 매트릭스·수수료·금리",
        "providerMatrices": matrices,
        "guaranteeFees": fees,
        "residualFloorByTerm": floors,
        "mileageAdj": {"10000": 0.0, "15000": 0.0, "20000": 0.0, "30000": -0.04},
        "mileageAdjSelf": {"10000": 0.0, "15000": 0.0, "20000": 0.0, "30000": -0.15},
        "brandRates": brand_rates,
        "rateSurcharges": {
            "lowDepositBelow11pct": 0.0015,
            "deposit40pctPlus": 0.003,
            "my24": 0.005,
        },
        "operatingFees": {
            "cmFeeRate": 0.03,
            "partnerFeeRate": 0.011,
            "extraFeeRate": 0.01,
            "fixedFee": 10000,
            "note": "CM 3%(입력 기본), 제휴사 1.1%(한성모터스), 추가 1%, 고정 1만",
        },
        "acquisition": {
            "taxRate": 0.07,
            "evTaxRebate": 1400000,
            "bondCost": 0,
            "deliveryCost": 0,
            "incidentalCost": 80000,
            "note": "공채(수원 0)·탁송 0 기본, 부대비 80,000 — 골든케이스 기준",
        },
        "financeLease": {"defaultRate": 0.063},
    }
    (OUT_DIR / "lease-data.json").write_text(json.dumps(data, ensure_ascii=False, indent=1))
    print("✅ vehicles.json / lease-data.json 저장")

    gk = vehicles.get("Benz E 220d 4MATIC AMG Line")
    print("골든케이스 차종:", gk)
    print("APS SA1:", matrices["aps"]["rates"].get("SA1"))


if __name__ == "__main__":
    if len(sys.argv) < 3 or sys.argv[1] != "extract":
        print(__doc__)
        sys.exit(1)
    extract(sys.argv[2])
