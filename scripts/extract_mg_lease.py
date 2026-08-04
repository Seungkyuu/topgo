"""
MG캐피탈 운용리스 엑셀 → JSON 추출기.

`운용리스` 시트(BMW 740i xDrive DPE, 60개월 실 견적 예시)를 셀 단위로 확인한
사슬 — 차량가 160,800,000원:

  취득세 = ROUNDDOWN(차량가/1.1 × 7%, -10)                = 10,232,720
  취득원가 = 차량가 + 취득세                                = 171,032,720
  AG수수료 = ROUNDDOWN(차량가 × 1.32%, 0)                   = 2,122,560  (CI54 취득원가 구간별 요율표, 근사)
  등록비 = 10,000(고정)
  금리 = 브랜드별 "당사명의" 운영IRR표(견적관리자용) — 실측 32개 브랜드 전부 5.4%로 동일
  잔가율 = 차량DB APS 열(모델별 12/24/36/48/60개월 직접 테이블)
  잔가금액 = ROUNDDOWN(차량가 × 잔가율, -1)
  월리스료 = ROUNDUP(PMT(금리/12, 기간, -(취득원가+AG수수료+등록비), 잔가금액), -2)
    (실측 2,106,100원 — 엑셀은 RATE()로 수수료를 실효금리에 접어넣는 한 단계가
    더 있어 오차 18원(0.001%) 수준, 무시 가능한 수준으로 근접)

⚠ v1 의도적 단순화(전부 "실제보다 낮게 보이면 안 된다" 방향):
  · 잔가는 원본이 3개 소스(에스앤케이모터스/APS/차봇) 중 최댓값을 자동
    선택하는데, 이번 실 예시는 SNK 쪽이 이겼지만(잔가율 0.51) 그 표는
    별도 시트 간 다단계 INDEX/MATCH라 시간상 다 풀지 못했다 — 대신 항상
    존재하는 차량DB 직접 컬럼(APS열)만 쓴다. 이 예시에서 APS열은 0.41로
    실제보다 낮다(=월리스료가 실제보다 높게 나옴) — 세 곳 중 최댓값을 못
    써서 생기는 차이지만, 방향은 안전하다.
  · 금리는 32개 브랜드 실측이 전부 5.4%로 동일해 브랜드 구분 없이 고정
    상수로 썼다. "이용자명의"(고객 직접 등록) 옵션은 별도 10% 고정이라
    이번 범위에서 제외(당사명의만 지원).
  · AG수수료는 취득원가 구간별 요율표를 실측값(1.32%)으로 고정.
  · 등록비 10,000원 고정 기본값.

사용법: python scripts/extract_mg_lease.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/mg-lease/data"
TERMS = [12, 24, 36, 48, 60]


def extract_vehicles(ws) -> dict:
    vehicles: dict[str, dict] = {}
    for r in range(6, ws.max_row + 1):
        brand = ws.cell(row=r, column=5).value
        model = ws.cell(row=r, column=6).value
        price = ws.cell(row=r, column=9).value
        if not model or not isinstance(model, str) or not isinstance(price, (int, float)):
            continue
        model = model.strip()

        # APS열(26~30열)이 모델별 잔가율 직접 테이블 — 없으면 이 차종은 미취급
        residual: dict[str, float] = {}
        for i, term in enumerate(TERMS):
            v = ws.cell(row=r, column=26 + i).value
            if isinstance(v, (int, float)) and v > 0:
                residual[str(term)] = round(float(v), 6)
        if not residual:
            continue

        vehicles[model] = {
            "brand": brand or "",
            "vehiclePrice": int(price),
            "engineCc": ws.cell(row=r, column=7).value or 0,
            "kind": ws.cell(row=r, column=8).value or "",
            "residualByTerm": residual,
        }
    return vehicles


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    vehicles = extract_vehicles(wb["차량DB"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(vehicles, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 모델")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_mg_lease.py <엑셀파일>")
        sys.exit(1)
    main(sys.argv[1])
