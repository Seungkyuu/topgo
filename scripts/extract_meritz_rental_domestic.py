"""
메리츠 국산차 장기렌트 엑셀 → JSON 추출기.

대표님이 전달한 실제 견적 예시(제네시스 GV70, 60개월/2만km/정비 Basic ->
월 859,540원)로 아래 수식 사슬을 셀 단위까지 정확히 검증했다(오차 0):

  차량가격(할인후) = 기본가격 + 옵션 - 최종할인 + 1차탁송료
  면세가격 = 공급가격A + 부가세A
    공급가격A = ROUND((기본가격+옵션)/개소세계수 + 1차탁송료/1.1, 0)
    부가세A  = ROUND(공급가격A × 10%, 0)
  출고가격 = 면세가격 - 최종할인
  공급가격B = ROUND(출고가격/1.1, 0)
  기준금액 = ROUNDDOWN(공급가격B - 1차탁송료/1.1, -3)
  개별소비세(현재기간 5.0%) = 기준금액×4%×130% - 감면(EV 390만/하이브리드 91만)
  개별소비세취득세 = ROUNDDOWN(개별소비세×4%, -1)
  취득세 = TRUNC(공급가격B×4% [- EV감면 140만], -1)
  취득원가 = 공급가격B + (2차탁송료+용품비)/1.1 + (개별소비세+개별소비세취득세)
             + 취득세 + 등록제비용 - (EV면 전기차보조금)

이 스크립트는 위 사슬에 필요한 모델별 참조데이터(잔가율표·전략등급·
정비비 Basic단가)만 추출한다 — 사슬 자체는 lib/engine/meritz-rental-domestic
/index.ts에 구현.

⚠ 개소세 5.0%/3.5% 구간은 정부 한시 정책(현재 "26년 7월 출고" 5.0% 구간)이라
언젠가 다시 바뀐다 — 엑셀이 갱신되면 이 스크립트도 같이 재실행해야 한다.

사용법: python scripts/extract_meritz_rental_domestic.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/meritz-rental-domestic/data"

TERM_MONTHS = [24, 36, 48, 60]
MILEAGE_BUCKETS = [10000, 15000, 20000, 25000, 30000, 40000, None]  # None=무제한

# 차량정보 시트: term 24개월 잔가율이 시작되는 컬럼(O=15)부터 term당 7컬럼
# (10k/15k/20k/25k/30k/40k/무제한), 24→36→48→60 순서로 이어짐(GV70 예시로 확인).
RESIDUAL_START_COL = 15  # O열


def col_letter(idx: int) -> str:
    return openpyxl.utils.get_column_letter(idx)


_BRAND_PREFIXES = ("제네시스 ", "현대 ", "기아 ")


def strip_brand_prefix(model: str) -> str:
    """원본 표기 불일치 정규화 — 차량정보/정비 두 시트 모두 트림에 따라
    브랜드 접두어가 붙었다 안 붙었다 한다(예: "GV70 EV AWD…"는 접두어 없음,
    "제네시스 GV70 2.5T 가솔린"은 있음). 이미 연결된 meritz-domestic(리스)
    카탈로그가 브랜드 접두어 없이 저장하므로 여기서도 항상 벗겨서 같은 차가
    두 개의 다른 모델로 갈라지지 않게 한다."""
    for prefix in _BRAND_PREFIXES:
        if model.startswith(prefix):
            return model[len(prefix):].strip()
    return model


import re

BRAND_DIVIDER_RE = re.compile(r"^■+\s*(.+?)\s*■+$")

# "차종" 헤더열만으로는 그 블록이 어느 브랜드인지 알 수 없다 — 실제 열 위치를
# 열어서 확인한 고정 매핑(2608.v1 기준). "일반"/"선구매프로모션" 블록은 고정
# 브랜드가 없고 내부에 "■■■■■■■■■■ 브랜드 ■■■■■■■■■■" 구분행으로 브랜드가
# 바뀌므로 이 매핑에 넣지 않는다(구분행을 실시간으로 추적해서 판별).
BLOCK_FIXED_BRAND = {
    74: "현대",
    144: "기아",
    214: "KGM",
    284: "르노",
    354: "쉐보레",
}


def find_blocks(ws) -> list[tuple[int, int]]:
    """'차량정보' 시트는 브랜드별로 나란히 놓인 여러 블록으로 구성된다
    (예: 일반/현대/기아/KG모빌리티/르노/쉐보레/선구매프로모션×2 — 국산 렌트 기준,
    수입 렌트는 일반/TESLA/폴스타/BYD). 각 블록은 자기 열 위치에 '차종' 헤더를
    갖는다 — 이 헤더 셀 좌표를 전부 찾아서 블록별로 독립적으로 순회한다.
    (예전엔 첫 블록(열 E)만 읽어서 브랜드 전용 블록에 있는 차종을 통째로
    누락시켰다 — 예: 국산 렌트의 제네시스 전 라인업이 '현대자동차' 블록에
    있는데 열 E만 읽어 전부 빠졌었다.)"""
    blocks: list[tuple[int, int]] = []
    for r in range(1, 10):
        for c in range(1, 600):
            if ws.cell(row=r, column=c).value == "차종":
                blocks.append((r, c))
    return blocks


def extract_vehicles(ws) -> dict:
    blocks = find_blocks(ws)
    if not blocks:
        raise RuntimeError("차량정보 헤더('차종') 못 찾음")

    vehicles: dict[str, dict] = {}
    for hdr_row, hdr_col in blocks:
        # "일반"/"선구매프로모션" 블록은 "■■■■■■■■■■ 브랜드 ■■■■■■■■■■" 구분행을
        # 만날 때마다 그 아래 차종들의 브랜드가 바뀐다 — 고정 브랜드 블록이면
        # 처음부터 그 브랜드로 시작.
        current_brand = BLOCK_FIXED_BRAND.get(hdr_col)
        for r in range(hdr_row + 1, hdr_row + 200):
            model = ws.cell(row=r, column=hdr_col).value
            if not model or not isinstance(model, str):
                continue
            model = model.strip()
            divider = BRAND_DIVIDER_RE.match(model)
            if divider:
                current_brand = divider.group(1).strip()
                continue
            model = strip_brand_prefix(model)
            if model.startswith("■") or model in ("", "-", "차종") or set(model) == {"-"}:
                continue

            fuel = ws.cell(row=r, column=hdr_col + 5).value  # 유종
            engine_cc = ws.cell(row=r, column=hdr_col + 7).value or 0  # 배기량
            strategy_grade = ws.cell(row=r, column=hdr_col + 4).value  # 전략구분
            kind = ws.cell(row=r, column=hdr_col + 2).value  # 차량등급
            insurance_grade = ws.cell(row=r, column=hdr_col + 3).value  # 보험등급(승합/경차 등 개소세 면제 판별용)
            tax_factor = ws.cell(row=r, column=hdr_col + 1).value  # 개소세계수(1+개소세율×1.3, 현재 정책 기준)

            residual: dict[str, float] = {}
            col = hdr_col + RESIDUAL_START_COL - 5
            for term in TERM_MONTHS:
                for mileage in MILEAGE_BUCKETS:
                    v = ws.cell(row=r, column=col).value
                    if isinstance(v, (int, float)) and mileage is not None:
                        residual[f"{term}_{mileage}"] = round(float(v), 4)
                    col += 1

            if not residual or not any(residual.values()):
                continue  # 잔가율 전부 0/미기입 — 구분선 등 비차량 행(예: "―――" 행)

            vehicles[model] = {
                "brand": current_brand or "",
                "fuel": fuel or "",
                "engineCc": int(engine_cc) if isinstance(engine_cc, (int, float)) else 0,
                "kind": kind or "",
                "insuranceGrade": insurance_grade or "",
                "strategyGrade": strategy_grade or "일반",
                "consumptionTaxFactor": round(float(tax_factor), 6)
                if isinstance(tax_factor, (int, float)) and tax_factor
                else 1.1572,
                "residualByTermMileage": residual,
            }
    return vehicles


def extract_maintenance(ws) -> dict:
    """정비서비스표준단가(Basic) 5개 컬럼(12/24/36/48/60개월)만 추출 —
    사용자 확정: 정비상품은 기본적으로 전부 Basic 등급으로 처리."""
    hdr_row = None
    for r in range(1, 5):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 25)]
        if any(v == "정비서비스표준단가(Basic)" for v in vals):
            hdr_row = r
            break
    if hdr_row is None:
        raise RuntimeError("정비 헤더('정비서비스표준단가(Basic)') 못 찾음")

    basic_start_col = None
    for c in range(1, 30):
        if ws.cell(row=hdr_row, column=c).value == "정비서비스표준단가(Basic)":
            basic_start_col = c
            break
    if basic_start_col is None:
        raise RuntimeError("Basic 단가 시작 컬럼 못 찾음")

    maintenance: dict[str, dict] = {}
    for r in range(hdr_row + 2, 400):
        model = ws.cell(row=r, column=3).value
        if not model or not isinstance(model, str):
            continue
        model = strip_brand_prefix(model.strip())
        vals = [ws.cell(row=r, column=basic_start_col + i).value for i in range(5)]
        if not all(isinstance(v, (int, float)) for v in vals):
            continue
        maintenance[model] = {
            "12": int(vals[0]),
            "24": int(vals[1]),
            "36": int(vals[2]),
            "48": int(vals[3]),
            "60": int(vals[4]),
        }
    return maintenance


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    vehicles = extract_vehicles(wb["차량정보"])
    maintenance = extract_maintenance(wb["정비"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(vehicles, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    (OUT_DIR / "maintenance.json").write_text(
        json.dumps(maintenance, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 모델")
    print(f"✅ maintenance.json: {len(maintenance)}개 모델")

    missing_maint = set(vehicles) - set(maintenance)
    if missing_maint:
        print(f"⚠ 정비 단가 없는 모델 {len(missing_maint)}개 (예: {list(missing_maint)[:5]})")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_meritz_rental_domestic.py <엑셀파일>")
        sys.exit(1)
    main(sys.argv[1])
