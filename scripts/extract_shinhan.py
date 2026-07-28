"""
신한카드 오토리스 엑셀 → JSON 추출기

사용법:
  python scripts/extract_shinhan.py extract <엑셀파일>   # vehicles/residual-groups/rates JSON 생성
  python scripts/extract_shinhan.py verify  <엑셀파일>   # 골든케이스 수식 검증만

─── 엑셀 수식 체인 (오토리스(운용&금융)_CA용 시트에서 해부, 2607_V1) ────────────
  AQ4  원금        = 차량가(H8/AD9) + 옵션(U8/AD10) - 할인(H9/AD11)
  AQ5  취득세      = ROUNDDOWN(AQ4/1.1 * 7%, -1) - 전기차감면(1,400,000)
  AQ7  기타        = 공채(AD16=80,000) + 탁송(AD15=168,000)
  AQ8  취득원가    = AQ4 + AQ5 + AQ6 + AQ7 + 인지대 5,000
  CA수수료(AN25)   = AQ8 * 수수료율(AD41, 골든케이스 4%)
  보증금(AN16)     = ROUNDUP(AQ4 * 보증금율, -4)          [40% 미만 구간]
  잔가(AN14)       = ROUNDUP(AQ4 * 잔가율, -4)
    잔가율 상한(AD25) = 잔가군표[모델.잔가군][기간] + 주행거리조정
    주행거리조정(AN15): 1만km +4% / 2만km +2% / 3만km 0% / 4만km -3% (표 기준 3만km)
    잔가율 하한: 12→54% 24→44% 36→34% 42→29% 44→29% 48→24% 60→20%
  월리스료(AR17)   = ROUNDUP(PMT((기준금리+0.01%)/12, 기간,
                       -(AQ8+CA수수료) + 보증금, 잔가 - 보증금), -1)
  기준금리 = 주요기준!C7 (2607: 5.9%)

─── 골든케이스 (시트 저장값 그대로) ─────────────────────────────────────────────
  포르쉐 CayenneCoupe가솔린3.0 (코드 450692470, 잔가군 R)
  차량가 148,400,000 + 옵션 21,000,000 - 할인 1,694,000 = 원금 167,706,000
  취득세 10,672,200 / 취득원가 178,631,200 / CA수수료 7,145,248
  60개월 / 잔가율 40%(선택) → 잔가 67,090,000 / 보증금 10% → 16,780,000
  → 월리스료 2,537,360원 ✅
"""

import sys
import json
import math
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl 후 재실행")
    sys.exit(1)

OUT_DIR = Path(__file__).parent.parent / "lib/engine/shinhan/data"

SHEET_MODELS = "모델, 잔가"       # 마스터 모델 테이블 (11행~, B브랜드 F모델 G코드 H가격 I배기량 J연료 M잔가군)
SHEET_GROUPS = "잔가군"           # 잔가군 매트릭스 (3행~, B코드, C:I = 12/24/36/42/44/48/60)
SHEET_BASE = "주요기준"           # C7 = 기준금리
GROUP_TERMS = ["12", "24", "36", "42", "44", "48", "60"]


def pmt(rate, nper, pv, fv=0):
    if rate == 0:
        return -(pv + fv) / nper
    f = (1 + rate) ** nper
    return -(pv * f + fv) * rate / (f - 1)


def roundup(v, digits):
    factor = 10 ** (-digits)
    return (math.ceil(v / factor) if v >= 0 else -math.ceil(-v / factor)) * factor


def rounddown(v, digits):
    factor = 10 ** (-digits)
    return (math.floor(v / factor) if v >= 0 else -math.floor(-v / factor)) * factor


def golden_check():
    """엑셀 저장값과 대조한 전체 수식 체인 검증."""
    principal = 148_400_000 + 21_000_000 - 1_694_000
    assert principal == 167_706_000
    tax = rounddown(principal / 1.1 * 0.07, -1)
    assert tax == 10_672_200, tax
    acq = principal + tax + 80_000 + 168_000 + 5_000
    assert acq == 178_631_200, acq
    fee = acq * 0.04
    assert fee == 7_145_248, fee
    deposit = roundup(principal * 0.10, -4)
    assert deposit == 16_780_000, deposit
    residual = roundup(principal * 0.40, -4)
    assert residual == 67_090_000, residual
    monthly = roundup(pmt((0.059 + 0.0001) / 12, 60, -(acq + fee) + deposit, residual - deposit), -1)
    assert monthly == 2_537_360, monthly
    print("✅ 골든케이스 수식 체인 검증 통과 (월리스료 2,537,360)")


def extract(xlsx_path: str):
    golden_check()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    # 1) 모델 마스터
    ws = wb[SHEET_MODELS]
    vehicles = {}
    dup_keys = set()
    for row in ws.iter_rows(min_row=11, max_row=700, min_col=1, max_col=14, values_only=True):
        brand, model, code, price, cc, fuel, group = row[1], row[5], row[6], row[7], row[8], row[9], row[12]
        if not brand or not model or not price or not group:
            continue
        key = f"{str(brand).strip()} {str(model).strip()}"
        if key in vehicles:  # 동명 모델(코드만 다름) → 코드 병기
            dup_keys.add(key)
            key = f"{key} ({code})"
        vehicles[key] = {
            "brand": str(brand).strip(),
            "model": str(model).strip(),
            "code": str(code),
            "vehiclePrice": int(price),
            "engineCc": int(cc or 0),
            "fuel": str(fuel).strip(),          # M=내연 / E=전기
            "residualGroup": str(group).strip(),
        }
    print(f"모델 {len(vehicles)}개 (중복명 {len(dup_keys)}개는 코드 병기)")

    # 2) 잔가군 매트릭스
    wsg = wb[SHEET_GROUPS]
    groups = {}
    for row in wsg.iter_rows(min_row=3, max_row=78, min_col=2, max_col=9, values_only=True):
        if row[0] is None:
            continue
        code = str(row[0]).strip()
        groups[code] = {t: round(float(row[i + 1]), 4) for i, t in enumerate(GROUP_TERMS) if row[i + 1] is not None}
    print(f"잔가군 {len(groups)}개")

    # 3) 기준금리
    base_rate = float(wb[SHEET_BASE]["C7"].value)
    print(f"기준금리 {base_rate}")

    # 사용된 잔가군이 모두 매트릭스에 있는지 확인
    used = {v["residualGroup"] for v in vehicles.values()}
    missing = used - set(groups)
    if missing:
        print(f"⚠️ 잔가군 매트릭스에 없는 코드: {missing}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    v = dict(vehicles)
    v["_note"] = "신한카드 오토리스(운용&금융) 2607_V1 엑셀 '모델, 잔가' 시트에서 추출"
    (OUT_DIR / "vehicles.json").write_text(json.dumps(v, ensure_ascii=False, indent=1))
    g = dict(groups)
    g["_note"] = "잔가군별 잔가율 (기간 12/24/36/42/44/48/60개월, 약정거리 3만km 기준)"
    (OUT_DIR / "residual-groups.json").write_text(json.dumps(g, ensure_ascii=False, indent=1))
    rates = {
        "operatingLease": {
            "baseRate": base_rate,
            "rateSurcharge": 0.0001,
            "note": "주요기준!C7 — 매달 변동. PMT에는 +0.01% 가산 적용(엑셀 AR17)",
        },
        "mileageResidualAdj": {"10000": 0.04, "20000": 0.02, "30000": 0.0, "40000": -0.03},
        "residualFloorByTerm": {"12": 0.54, "24": 0.44, "36": 0.34, "42": 0.29, "44": 0.29, "48": 0.24, "60": 0.20},
        "acquisition": {
            "taxRate": 0.07,
            "evTaxRebate": 1400000,
            "bondCost": 80000,
            "deliveryCost": 168000,
            "stampDuty": 5000,
            "note": "취득원가 = 원금 + ROUNDDOWN(원금/1.1*7%,-1) + 공채 + 탁송 + 인지대. 공채/탁송은 골든케이스 기본값(견적별 변동 가능)",
        },
        "defaultCaFeeRate": 0.04,
    }
    (OUT_DIR / "rates.json").write_text(json.dumps(rates, ensure_ascii=False, indent=1))
    print("✅ vehicles.json / residual-groups.json / rates.json 저장")

    # 골든케이스 데이터 정합성
    gk = vehicles.get("포르쉐 CayenneCoupe가솔린3.0")
    if gk:
        ok = gk["vehiclePrice"] == 148_400_000 and gk["residualGroup"] == "R" and groups["R"]["60"] == 0.39
        print(f"골든케이스 데이터 정합성: {'✅' if ok else '❌'} {gk}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == "verify":
        golden_check()
    elif cmd == "extract":
        extract(sys.argv[2])
    else:
        print(f"알 수 없는 명령: {cmd}  (extract | verify)")
        sys.exit(1)
