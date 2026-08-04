"""
MG캐피탈 장기렌터카 엑셀 → JSON 추출기.

`견적서및입력시트`(1164행×133열, PV5_패신저 실 견적 예시)를 셀 단위로 확인한
사슬 — 기본가 52,860,000원·36개월·연 2만km:

  면세가A(공급) = ROUND(기본가/특소세계수 + 탁송비/1.1, 0)          = 45,679,226
  면세가A(부가세) = ROUND(공급×10%, 0)                             = 4,567,923
  출고가(세금계산서) = 면세가A(공급+부가세) − 할인                  = 50,247,149
  공급가액(개소세 베이스) = ROUNDDOWN(출고가/1.1, 0)                = 44,537,273
  개별소비세 = ROUNDDOWN(IF(EV, MAX(공급가액×4%−140만,0), 공급가액×4%), -1) = 381,490
  취득원가 = 공급가액 + 등록및탁송비용 + 개별소비세 − EV보조금/1.1  = 39,279,444.8
  잔가율 = 기본잔가(잔가 시트, 잔가군×기간) + 약정거리조정 + 차종특별잔가(24개월초과만)
         = 0.42 + 0.01 + 0.04 = 0.47
  잔가금액 = ROUNDDOWN(차량총금액 × 잔가율, -1)                    = 24,844,200
  금리 = 운영기준 시트, 모델×기간                                   = 0.052
  기본대여료 = ROUNDUP(PMT(금리/12, 기간, −취득원가, 잔가금액/1.1), -2) (오차 0)

⚠ v1 의도적 단순화(전부 "실제보다 낮게 보이면 안 된다" 방향):
  · 잔가보장사_잔가의 "4만/5만pro·event Promotion·약정주행프로모션_②·
    1.7만 프로모션·48/60개월 추가잔가"는 특정 약정거리 조합에서만 켜지는
    프로모션이라 전부 미반영(0) — "차종 특별 잔가"(24개월초과 시 전 차종
    공통 적용)만 반영한다.
  · 전기차보조금은 지역·시기별 수기입력값이라 카탈로그에 없음 — 입력
    기본값 0.
  · 자동차세·보험료는 차종별 정확한 산정식(배기량 등급표·다요소 보험료
    산식)을 다 풀지 못해, PV5_패신저 실 견적값(월 65,700원·99,050원)을
    안전한 근사 기본값으로 쓴다 — 실제보다 낮게 보이는 위험보단, 상담에서
    실제 차종 기준으로 조정될 여지를 남기는 쪽이 안전하다.
  · 등록및탁송비용(DT5 합계, 360,682원)도 지역·차종별 세부 항목 대신
    검증된 예시의 합계값을 고정 기본값으로 사용.

사용법: python scripts/extract_mg_rental.py <엑셀파일>
"""

import json
import sys
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).parent.parent / "lib/engine/mg-rental/data"
TERMS = [12, 24, 36, 48, 60]


def extract_vehicles(wb) -> dict:
    ws_list = wb["차량_List"]
    ws_guarantor = wb["잔가보장사_잔가"]

    # 잔가보장사_잔가: 모델명 → {잔가군 코드, 차종특별잔가}
    guarantor_by_model: dict[str, dict] = {}
    for r in range(4, ws_guarantor.max_row + 1):
        model = ws_guarantor.cell(row=r, column=3).value
        residual_class = ws_guarantor.cell(row=r, column=4).value
        special_bonus = ws_guarantor.cell(row=r, column=8).value
        if not model or not isinstance(model, str) or residual_class is None:
            continue
        guarantor_by_model[model.strip()] = {
            "residualClass": int(residual_class),
            "specialResidualBonus": round(float(special_bonus), 6)
            if isinstance(special_bonus, (int, float))
            else 0,
        }

    vehicles: dict[str, dict] = {}
    for r in range(4, ws_list.max_row + 1):
        brand = ws_list.cell(row=r, column=2).value
        model = ws_list.cell(row=r, column=3).value
        if not model or not isinstance(model, str):
            continue
        model = model.strip()
        fuel = ws_list.cell(row=r, column=5).value
        tax_factor = ws_list.cell(row=r, column=6).value
        insurance_grade = ws_list.cell(row=r, column=8).value

        g = guarantor_by_model.get(model)
        if g is None:
            continue  # 잔가보장사 매칭 없는 차종은 잔가율 산정 불가 — 미취급

        vehicles[model] = {
            "brand": brand or "",
            "fuel": fuel or "",
            "consumptionTaxFactor": round(float(tax_factor), 6)
            if isinstance(tax_factor, (int, float)) and tax_factor
            else 1.1572,
            "insuranceGrade": insurance_grade or "",
            "residualClass": g["residualClass"],
            "specialResidualBonus": g["specialResidualBonus"],
        }
    return vehicles


def extract_residual_matrix(ws) -> dict:
    """`잔가` 시트 "□ 무카 잔가군" 표(행 65 헤더, 66~ 데이터) — 잔가군(정수) × 기간(12/24/36/48/60)."""
    matrix: dict[str, dict[str, float]] = {}
    for r in range(66, ws.max_row + 1):
        cls = ws.cell(row=r, column=2).value
        if not isinstance(cls, (int, float)):
            continue
        row_vals: dict[str, float] = {}
        for i, term in enumerate(TERMS):
            v = ws.cell(row=r, column=3 + i).value
            if isinstance(v, (int, float)):
                row_vals[str(term)] = round(float(v), 6)
        if row_vals:
            matrix[str(int(cls))] = row_vals
    return matrix


def extract_mileage_adjustment(ws) -> dict:
    """견적서및입력시트 CV101:CW110 — 연간 약정주행거리 → 잔가율 가감."""
    result: dict[str, float] = {}
    col_km, col_adj = 100, 101  # CV, CW (1-based)
    for r in range(101, 111):
        km = ws.cell(row=r, column=col_km).value
        adj = ws.cell(row=r, column=col_adj).value
        if isinstance(km, (int, float)) and isinstance(adj, (int, float)):
            result[str(int(km))] = round(float(adj), 6)
    return result


def extract_rate_table(ws) -> dict:
    """`운영기준` 시트 — 모델별 제조사할인율 + 기간(24/36/48/60)별 적용금리."""
    rates: dict[str, dict] = {}
    for r in range(4, ws.max_row + 1):
        model = ws.cell(row=r, column=3).value
        if not model or not isinstance(model, str):
            continue
        by_term: dict[str, float] = {}
        for i, term in enumerate([24, 36, 48, 60]):
            v = ws.cell(row=r, column=6 + i).value
            if isinstance(v, (int, float)):
                by_term[str(term)] = round(float(v), 6)
        if not by_term:
            continue
        discount = ws.cell(row=r, column=4).value  # 제조사할인율
        rates[model.strip()] = {
            "byTerm": by_term,
            "manufacturerDiscountRate": round(float(discount), 6)
            if isinstance(discount, (int, float))
            else 0,
        }
    return rates


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    vehicles = extract_vehicles(wb)
    residual_matrix = extract_residual_matrix(wb["잔가"])
    mileage_adjustment = extract_mileage_adjustment(wb["견적서및입력시트"])
    rate_table = extract_rate_table(wb["운영기준"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "vehicles.json").write_text(
        json.dumps(vehicles, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    (OUT_DIR / "residual-matrix.json").write_text(
        json.dumps(residual_matrix, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    (OUT_DIR / "mileage-adjustment.json").write_text(
        json.dumps(mileage_adjustment, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    (OUT_DIR / "rate-table.json").write_text(
        json.dumps(rate_table, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    )
    print(f"✅ vehicles.json: {len(vehicles)}개 모델")
    print(f"✅ residual-matrix.json: {len(residual_matrix)}개 잔가군")
    print(f"✅ mileage-adjustment.json: {mileage_adjustment}")
    print(f"✅ rate-table.json: {len(rate_table)}개 모델")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("사용법: python scripts/extract_mg_rental.py <엑셀파일>")
        sys.exit(1)
    main(sys.argv[1])
